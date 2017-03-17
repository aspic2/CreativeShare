#MT: Handles reading and writing Excel spreadsheet info

from openpyxl import load_workbook
from openpyxl import Workbook
from os import getcwd

'''To make a new workbook, set the second parameter 'assemble=' to True
To use an existing workbook, include a path.
'''
class Spreadsheet(object):
    #name will be used for the saved file
    def __init__(self, name='test', assemble=False, path=None):
        self.name = name
        self.assemble = assemble
        self.path = path
        if self.assemble:
            self.wb = Workbook(write_only=True)
            self.ws = self.wb.create_sheet()
            #match these with values of token in zombiescript.main()
            title = ("NewLID", 'Name', 'Status', 'Errors')
            self.ws.append(title)
            self.assemble = False

    def read(self):
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        self.values = []
        for row in ws.rows:
            lineinfo = (row[0].value, row[1].value, row[2].value)
            self.values.append(lineinfo)
            #row = (a.value, b.value, c.value)
            #try:
                #int(a.value)
                #int(c.value)
            #self.values.append(row)
            #except Exception as ValueError:
                #continue
        return self.values

    def write_wb(self, values):
        self.ws.append(values)

    def save(self):
        file_name = str(getcwd() + '\\Output\\' + self.name + '.xlsx')
        self.wb.save(filename=file_name)
